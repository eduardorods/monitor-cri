"""Imprime o conteúdo de um arquivo .xlsx no console (uso em logs de CI)."""
import sys

import openpyxl


def main(path: str) -> int:
    wb = openpyxl.load_workbook(path, data_only=True)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        print(f"\n=== {sheet} ===")
        for row in ws.iter_rows(values_only=True):
            print("\t".join(str(c or "") for c in row))
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1]))
