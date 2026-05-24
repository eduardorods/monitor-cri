import argparse
import json
import os
import sys
from dataclasses import asdict
from typing import Optional

from cri_monitor import B3Client
from cri_monitor.b3 import eh_consolidado
from cri_monitor.drive import baixar_manuais
from cri_monitor.fnet import baixar_documento
from cri_monitor.pdf_parser import ResumoOperacao, analisar, extrair_texto
from cri_monitor.gemini_parser import analisar_ata_com_gemini, analisar_com_gemini


CATEGORIAS_BAIXAR_DEFAULT = (
    "termo_securitizacao",
    "aditamento",
    "ata_assembleia",
    "relatorio_agente_fiduciario",
)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Consulta dados públicos de CRI na API da B3.",
    )
    parser.add_argument("--codigo-if",
                        help="Código IF do CRI (ex.: 19F0076447).")
    parser.add_argument("--cnpj-securitizadora",
                        help="CNPJ da securitizadora (só dígitos).")
    parser.add_argument("--sem-documentos", action="store_true",
                        help="Não buscar documentos FNet.")
    parser.add_argument("--saida",
                        help="Arquivo de saída (.json ou .xlsx). Se omitido, imprime no stdout.")
    parser.add_argument("--baixar-pdfs", metavar="DIR",
                        help="Diretório onde salvar os PDFs baixados (e parsear o Termo).")

    parser.add_argument("--listar-securitizadoras", action="store_true",
                        help="Dump bruto das securitizadoras retornadas pela B3.")
    parser.add_argument("--listar-cris", action="store_true",
                        help="Dump bruto dos CRIs da securitizadora "
                             "(exige --cnpj-securitizadora).")
    parser.add_argument("--limite", type=int, default=0,
                        help="Limite de itens no modo listar (0 = todos).")

    args = parser.parse_args()
    client = B3Client()

    if args.listar_securitizadoras:
        print("Listando securitizadoras (API B3 + lista local)...", file=sys.stderr)
        from cri_monitor.securitizadoras import SECURITIZADORAS_CONHECIDAS
        from cri_monitor.b3 import _merge_securitizadoras
        merged = _merge_securitizadoras(list(client.securitizadoras()), SECURITIZADORAS_CONHECIDAS)
        print(f"{len(merged)} securitizadora(s) no total.", file=sys.stderr)
        return _dump(merged, args.saida)

    if args.listar_cris:
        if not args.cnpj_securitizadora:
            parser.error("--listar-cris exige --cnpj-securitizadora")
        cnpj = _normalizar_cnpj(args.cnpj_securitizadora)
        print(f"Listando CRIs do CNPJ {cnpj}...", file=sys.stderr)
        cris = []
        for cri in client.cris_por_securitizadora(cnpj):
            cris.append(cri)
            if args.limite and len(cris) >= args.limite:
                break
        print(f"{len(cris)} CRI(s) retornado(s).", file=sys.stderr)
        return _dump(cris, args.saida)

    if not args.codigo_if:
        parser.error("informe --codigo-if (ou use --listar-securitizadoras "
                     "/ --listar-cris para debug).")

    cnpj = _normalizar_cnpj(args.cnpj_securitizadora) if args.cnpj_securitizadora else None
    print(f"Buscando CRI {args.codigo_if}...", file=sys.stderr)

    def _progress(idx, total, nome):
        print(f"  [{idx}/{total}] {nome}", file=sys.stderr)

    info = client.buscar_cri(
        codigo_if=args.codigo_if,
        cnpj_securitizadora=cnpj,
        incluir_documentos=not args.sem_documentos,
        on_progress=_progress if not cnpj else None,
    )
    if info is None:
        print(f"CRI {args.codigo_if} não encontrado.", file=sys.stderr)
        return 1

    resumo: Optional[ResumoOperacao] = None
    pdfs_baixados: list[str] = []
    atas_analisadas: list[dict] = []
    termo_origem: str = "n/a"

    if args.baixar_pdfs and info.documentos:
        pdfs_baixados, resumo, atas_analisadas, termo_origem = _baixar_e_analisar(
            info, args.baixar_pdfs)
        _preencher_lacunas(info, resumo)

    payload = info.to_dict()
    payload["termo_origem"] = termo_origem
    if resumo is not None:
        payload["resumo_operacao"] = asdict(resumo)
    if pdfs_baixados:
        payload["pdfs_baixados"] = pdfs_baixados
    if atas_analisadas:
        payload["atas_analisadas"] = atas_analisadas

    return _dump(payload, args.saida)


def _selecionar_docs_termo(documentos):
    """Decide quais documentos (termo + aditamentos) devem ser analisados juntos.

    Regra:
    1. Se existir versão CONSOLIDADA (termo ou aditamento com "consolidado" no nome),
       usa apenas a mais recente — ela já incorpora os aditamentos.
    2. Caso contrário, usa o termo ORIGINAL (mais antigo) + TODOS os aditamentos
       em ordem cronológica (mais antigo → mais recente), para o Gemini conseguir
       aplicar as emendas em ordem.
    3. Se não houver termo original, usa apenas aditamentos (sinal de falha de
       indexação, mas é o melhor que temos).

    Retorna: lista de Documento(s) na ordem em que devem aparecer na análise.
    """
    termos = [d for d in documentos if d.categoria_normalizada == "termo_securitizacao"]
    aditamentos = [d for d in documentos if d.categoria_normalizada == "aditamento"]

    consolidados = [d for d in (termos + aditamentos) if eh_consolidado(d.nome, d.tipo)]
    if consolidados:
        # info.documentos vem ordenado por dataEntrega desc → primeiro = mais recente
        return [consolidados[0]]

    selecionados = []
    if termos:
        # Mais antigo = última posição na lista desc
        selecionados.append(termos[-1])
    # Aditamentos em ordem cronológica (do mais antigo ao mais recente)
    selecionados.extend(reversed(aditamentos))
    return selecionados


def _nome_base_doc(doc, idx_fallback: int) -> str:
    """Nome de arquivo determinístico para um Documento FNet.

    Usa o ID do FNet como prefixo (estável entre execuções → dedupe funciona).
    Fallback: usa índice posicional (raro, só se a fonte não traz ID).
    """
    prefixo = f"fnet{doc.fnet_id}" if doc.fnet_id else f"i{idx_fallback:03d}"
    cat = doc.categoria_normalizada or "doc"
    nome = doc.nome or "doc"
    return f"{prefixo}_{cat}_{nome}"


def _baixar_e_analisar(info, dir_destino: str):
    """Baixa PDFs relevantes e analisa Termo (+ aditamentos quando aplicável).

    Ordem de prioridade para o "termo principal":
    1. Arquivo `manual_*` no Drive (escape hatch quando FNet não tem o original).
    2. Termo FNet (ou versão consolidada, se existir).
    3. Apenas aditamentos (degraded).

    Retorna (pdfs_baixados, resumo, atas_analisadas, termo_origem).
    termo_origem ∈ {'manual_drive', 'fnet', 'fnet_consolidado',
                    'fnet_so_aditamentos', 'ausente'}.
    """
    LIMITE_CATEGORIA = {
        "ata_assembleia": 5,
        "relatorio_agente_fiduciario": 3,
    }
    alvo_dir = os.path.join(dir_destino, info.codigo_if or "cri")
    print(f"Baixando PDFs para {alvo_dir}...", file=sys.stderr)
    print(f"  Total de documentos disponíveis: {len(info.documentos)}", file=sys.stderr)

    # 1) Tenta detectar arquivo manual no Drive
    manual_termo = None
    subpasta_drive = f"CRIs/{info.codigo_if}" if info.codigo_if else None
    if subpasta_drive:
        manuais = baixar_manuais(subpasta_drive, alvo_dir)
        for m in manuais:
            nome_min = os.path.basename(m).lower()
            if "termo" in nome_min:
                manual_termo = m
                break
        if manuais:
            print(f"  Drive: {len(manuais)} arquivo(s) manual(is) detectado(s); "
                  f"manual_termo={'sim' if manual_termo else 'não'}", file=sys.stderr)

    # 2) Seleciona termo + aditamentos do FNet
    docs_termo_fnet = _selecionar_docs_termo(info.documentos)
    ids_termo = {id(d) for d in docs_termo_fnet}
    tem_termo_fnet = any(d.categoria_normalizada == "termo_securitizacao"
                          for d in docs_termo_fnet)
    tem_consolidado = any(eh_consolidado(d.nome, d.tipo) for d in docs_termo_fnet)

    if docs_termo_fnet:
        print(f"  FNet: {len(docs_termo_fnet)} documento(s) selecionado(s):",
              file=sys.stderr)
        for d in docs_termo_fnet:
            marca = " [CONSOLIDADO]" if eh_consolidado(d.nome, d.tipo) else ""
            print(f"    - [{d.categoria_normalizada}] {d.nome or '(sem nome)'} "
                  f"({d.data_entrega or 's/d'}){marca}", file=sys.stderr)

    # 3) Decide a origem do termo principal
    if manual_termo:
        termo_origem = "manual_drive"
    elif tem_consolidado:
        termo_origem = "fnet_consolidado"
    elif tem_termo_fnet:
        termo_origem = "fnet"
    elif docs_termo_fnet:  # só aditamentos
        termo_origem = "fnet_so_aditamentos"
    else:
        termo_origem = "ausente"
    print(f"  Origem do termo principal: {termo_origem}", file=sys.stderr)
    if termo_origem == "ausente":
        print(f"  AVISO: nenhum termo de securitização encontrado no FNet nem manual "
              f"no Drive. Para corrigir, faça upload do PDF em "
              f"Drive:/CRIs/{info.codigo_if}/ com nome iniciado por 'manual_termo' "
              f"e re-execute.", file=sys.stderr)
    elif termo_origem == "fnet_so_aditamentos":
        print(f"  AVISO: FNet só tem aditamentos (sem termo original). Análise pode "
              f"ficar incompleta. Para análise completa, faça upload do termo "
              f"original em Drive:/CRIs/{info.codigo_if}/manual_termo.pdf.",
              file=sys.stderr)

    # 4) Download dos docs do FNet (com dedupe por nome local determinístico)
    contagem: dict[str, int] = {}
    pdfs_baixados: list[str] = []
    caminhos_termo_fnet: list[tuple[str, object]] = []
    caminhos_atas: list[str] = []

    for i, doc in enumerate(info.documentos, 1):
        cat = doc.categoria_normalizada
        eh_doc_termo = id(doc) in ids_termo
        if cat not in CATEGORIAS_BAIXAR_DEFAULT:
            continue
        if not doc.url:
            continue
        if not eh_doc_termo and cat in LIMITE_CATEGORIA:
            if contagem.get(cat, 0) >= LIMITE_CATEGORIA[cat]:
                continue
            contagem[cat] = contagem.get(cat, 0) + 1
        elif not eh_doc_termo and cat in ("termo_securitizacao", "aditamento"):
            continue

        nome_base = _nome_base_doc(doc, i)
        caminho = baixar_documento(doc.url, alvo_dir, nome_base)
        if caminho:
            print(f"  [{cat}] Baixado: {os.path.basename(caminho)}", file=sys.stderr)
            pdfs_baixados.append(caminho)
            if eh_doc_termo:
                caminhos_termo_fnet.append((caminho, doc))
            elif cat == "ata_assembleia":
                caminhos_atas.append(caminho)

    # Mantém ordem cronológica (consolidado primeiro, ou original → aditamentos)
    ordem = {id(d): i for i, d in enumerate(docs_termo_fnet)}
    caminhos_termo_fnet.sort(key=lambda cd: ordem.get(id(cd[1]), 999))

    # 5) Monta a lista final de docs para análise
    partes: list[tuple[str, str]] = []  # (cabeçalho, texto)
    if manual_termo:
        print(f"  Extraindo texto (manual): {os.path.basename(manual_termo)}",
              file=sys.stderr)
        partes.append((f"TERMO DE SECURITIZAÇÃO (UPLOAD MANUAL) — "
                       f"{os.path.basename(manual_termo)}",
                       extrair_texto(manual_termo)))
        # Quando temos manual, ignoramos o termo original do FNet (manual o substitui)
        # mas mantemos os aditamentos do FNet
        for caminho, doc in caminhos_termo_fnet:
            if doc.categoria_normalizada == "aditamento":
                print(f"  Extraindo texto (FNet): {os.path.basename(caminho)}",
                      file=sys.stderr)
                cab = (f"{doc.categoria_normalizada.upper()} — {doc.nome or ''} "
                       f"({doc.data_entrega or 's/d'})")
                partes.append((cab, extrair_texto(caminho)))
    else:
        for caminho, doc in caminhos_termo_fnet:
            print(f"  Extraindo texto (FNet): {os.path.basename(caminho)}",
                  file=sys.stderr)
            cab = (f"{doc.categoria_normalizada.upper()} — {doc.nome or ''} "
                   f"({doc.data_entrega or 's/d'})")
            partes.append((cab, extrair_texto(caminho)))

    # 6) Análise
    resumo: Optional[ResumoOperacao] = None
    if partes:
        if len(partes) == 1:
            texto_combinado = partes[0][1]
            multiplos = False
        else:
            texto_combinado = "\n\n".join(
                f"=== DOCUMENTO {i+1}/{len(partes)}: {cab} ===\n\n{txt}"
                for i, (cab, txt) in enumerate(partes)
            )
            multiplos = True

        if os.environ.get("GOOGLE_API_KEY"):
            print(f"  Tentando análise com Gemini ({len(partes)} doc(s), "
                  f"{len(texto_combinado):,} chars)...", file=sys.stderr)
            resumo = analisar_com_gemini(texto_combinado, multiplos_docs=multiplos)
            if resumo:
                print("  Análise concluída via Gemini.", file=sys.stderr)
        if resumo is None:
            print("  Análise via regex (sem GOOGLE_API_KEY ou Gemini indisponível).",
                  file=sys.stderr)
            resumo = analisar(texto_combinado)

    atas_analisadas: list[dict] = []
    if os.environ.get("GOOGLE_API_KEY") and caminhos_atas:
        print(f"Analisando {len(caminhos_atas)} ata(s) com Gemini...", file=sys.stderr)
        for caminho_ata in caminhos_atas:
            nome = os.path.basename(caminho_ata)
            print(f"  Ata: {nome}", file=sys.stderr)
            texto_ata = extrair_texto(caminho_ata, max_paginas=30)
            resultado = analisar_ata_com_gemini(texto_ata, nome)
            if resultado:
                atas_analisadas.append(resultado)

    return pdfs_baixados, resumo, atas_analisadas, termo_origem


def _preencher_lacunas(info, resumo: Optional[ResumoOperacao]) -> None:
    if not resumo:
        return
    if not info.devedor and resumo.devedor:
        info.devedor = resumo.devedor
    if not info.data_emissao and resumo.data_emissao:
        info.data_emissao = resumo.data_emissao
    if not info.data_vencimento and resumo.data_vencimento:
        info.data_vencimento = resumo.data_vencimento
    if not info.remuneracao and resumo.taxa_remuneracao:
        info.remuneracao = resumo.taxa_remuneracao


def _dump(obj, saida) -> int:
    if saida and saida.lower().endswith(".xlsx"):
        _dump_excel(obj, saida)
        print(f"Resultado salvo em {saida}", file=sys.stderr)
        return 0
    payload = json.dumps(obj, ensure_ascii=False, indent=2, default=str)
    if saida:
        with open(saida, "w", encoding="utf-8") as f:
            f.write(payload)
        print(f"Resultado salvo em {saida}", file=sys.stderr)
    else:
        print(payload)
    return 0


def _dump_excel(obj, path: str) -> None:
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    if isinstance(obj, dict):
        # Aba Resumo
        ws = wb.create_sheet("Resumo")
        ORIGEM_LABEL = {
            "manual_drive":         "Upload manual (Drive)",
            "fnet_consolidado":     "FNet (versão consolidada)",
            "fnet":                 "FNet (termo + aditamentos)",
            "fnet_so_aditamentos":  "FNet (SÓ aditamentos — análise incompleta)",
            "ausente":              "AUSENTE — faça upload em Drive:/CRIs/{cod}/manual_termo.pdf",
            "n/a":                  "—",
        }
        origem = obj.get("termo_origem") or "n/a"
        origem_label = ORIGEM_LABEL.get(origem, origem)
        if origem == "ausente":
            origem_label = origem_label.replace("{cod}", obj.get("codigo_if") or "<código>")

        campos = [
            ("Código IF",          "codigo_if"),
            ("ISIN",               "isin"),
            ("Securitizadora",     "nome_securitizadora"),
            ("CNPJ Securitizadora","cnpj_securitizadora"),
            ("Agente Fiduciário",  "agente_fiduciario"),
            ("Emissão",            "emissao"),
            ("Série",              "serie"),
            ("Data de Emissão",    "data_emissao"),
            ("Data de Vencimento", "data_vencimento"),
            ("Remuneração",        "remuneracao"),
            ("Devedor",            "devedor"),
            ("Descrição",          "descricao"),
            ("Fase",               "fase"),
            ("Origem do Termo",    None),  # tratado abaixo
        ]
        for i, (label, key) in enumerate(campos, 1):
            ws.cell(row=i, column=1, value=label).font = Font(bold=True)
            if key is None and label == "Origem do Termo":
                ws.cell(row=i, column=2, value=origem_label)
            else:
                ws.cell(row=i, column=2, value=obj.get(key))
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 55

        # Aba Resumo da Operação (parsing do Termo de Securitização)
        resumo = obj.get("resumo_operacao")
        if resumo:
            wo = wb.create_sheet("Resumo da Operação")
            ordem = [
                ("Devedor",                   "devedor"),
                ("CNPJ do Devedor",           "cnpj_devedor"),
                ("Valor Total",               "valor_total"),
                ("Data de Emissão",           "data_emissao"),
                ("Data de Vencimento",        "data_vencimento"),
                ("Taxa / Remuneração",        "taxa_remuneracao"),
                ("Índice",                    "indice_atualizacao"),
                ("Coordenador Líder",         "coordenador_lider"),
                ("Agente Fiduciário",         "agente_fiduciario"),
                ("Lastro",                    "lastro"),
                ("Cronograma de Pagamentos",  "cronograma_descricao"),
                ("Garantias (resumo)",        None),
                ("Garantias (detalhadas)",    "garantias_detalhadas"),
            ]
            row_idx = 1
            for label, key in ordem:
                wo.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
                if key == "garantias_detalhadas":
                    cell = wo.cell(row=row_idx, column=2, value=resumo.get(key))
                    cell.alignment = __import__("openpyxl").styles.Alignment(wrap_text=True)
                elif key == "cronograma_descricao":
                    cell = wo.cell(row=row_idx, column=2, value=resumo.get(key))
                    cell.alignment = __import__("openpyxl").styles.Alignment(wrap_text=True)
                elif key is None:
                    garantias = resumo.get("garantias") or []
                    wo.cell(row=row_idx, column=2,
                            value=", ".join(garantias) if garantias else None)
                else:
                    wo.cell(row=row_idx, column=2, value=resumo.get(key))
                row_idx += 1

            covenants = resumo.get("covenants") or []
            if covenants:
                wo.cell(row=row_idx, column=1, value="Covenants").font = Font(bold=True)
                cell = wo.cell(row=row_idx, column=2, value="\n".join(covenants))
                cell.alignment = __import__("openpyxl").styles.Alignment(wrap_text=True)
                row_idx += 1

            serie_unica = resumo.get("serie_unica")
            series = resumo.get("series") or []
            if serie_unica is not None or series:
                wo.cell(row=row_idx, column=1, value="Série Única?").font = Font(bold=True)
                wo.cell(row=row_idx, column=2,
                        value="Sim" if serie_unica else ("Não" if serie_unica is False else None))
                row_idx += 1
                for s in series:
                    if not isinstance(s, dict):
                        continue
                    label = s.get("serie", "Série")
                    partes = []
                    if s.get("valor"):
                        partes.append(f"Valor: {s['valor']}")
                    if s.get("quantidade_titulos"):
                        partes.append(f"Qtd: {s['quantidade_titulos']}")
                    if s.get("remuneracao"):
                        partes.append(f"Remuneração: {s['remuneracao']}")
                    if s.get("data_vencimento"):
                        partes.append(f"Vencimento: {s['data_vencimento']}")
                    if s.get("descricao"):
                        partes.append(s["descricao"])
                    wo.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
                    cell = wo.cell(row=row_idx, column=2, value=" | ".join(partes))
                    cell.alignment = __import__("openpyxl").styles.Alignment(wrap_text=True)
                    row_idx += 1

            wo.column_dimensions["A"].width = 28
            wo.column_dimensions["B"].width = 80
            for r in range(1, row_idx + 2):
                wo.row_dimensions[r].height = None  # auto

        # Aba Cronograma (tabela de pagamentos extraída do Termo)
        tabela = (resumo or {}).get("tabela_cronograma") or []
        if tabela and isinstance(tabela, list) and len(tabela) > 0:
            wc = wb.create_sheet("Cronograma")
            all_keys = []
            for row in tabela:
                if isinstance(row, dict):
                    for k in row.keys():
                        if k not in all_keys:
                            all_keys.append(k)
            header_labels = {
                "data": "Data",
                "evento": "Evento",
                "percentual_amortizacao": "Amortização %",
                "valor_parcela": "Valor Parcela",
                "saldo_devedor": "Saldo Devedor",
                "observacao": "Observação",
            }
            headers = [header_labels.get(k, k) for k in all_keys]
            for col, h in enumerate(headers, 1):
                wc.cell(row=1, column=col, value=h).font = Font(bold=True)
            for row_num, row_data in enumerate(tabela, 2):
                if isinstance(row_data, dict):
                    for col, k in enumerate(all_keys, 1):
                        wc.cell(row=row_num, column=col, value=row_data.get(k))
            for col in range(1, len(headers) + 1):
                wc.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

        # Aba Atas de Assembleia
        atas = obj.get("atas_analisadas") or []
        if atas:
            wa = wb.create_sheet("Atas de Assembleia")
            headers_atas = ["Data", "Tipo", "Arquivo", "Deliberações"]
            keys_atas = ["data_assembleia", "tipo_assembleia", "arquivo", "deliberacoes"]
            for col, h in enumerate(headers_atas, 1):
                wa.cell(row=1, column=col, value=h).font = Font(bold=True)
            for row_num, ata in enumerate(atas, 2):
                for col, k in enumerate(keys_atas, 1):
                    cell = wa.cell(row=row_num, column=col, value=ata.get(k))
                    if k == "deliberacoes":
                        cell.alignment = __import__("openpyxl").styles.Alignment(wrap_text=True)
                        wa.row_dimensions[row_num].height = 80
            wa.column_dimensions["A"].width = 15
            wa.column_dimensions["B"].width = 30
            wa.column_dimensions["C"].width = 35
            wa.column_dimensions["D"].width = 80

        # Aba Documentos
        docs = obj.get("documentos") or []
        if docs:
            wd = wb.create_sheet("Documentos")
            headers = ["Nome", "Tipo", "Categoria", "Categoria Norm.",
                       "Data Entrega", "Data Referência", "URL"]
            keys    = ["nome", "tipo", "categoria", "categoria_normalizada",
                       "data_entrega", "data_referencia", "url"]
            for col, h in enumerate(headers, 1):
                wd.cell(row=1, column=col, value=h).font = Font(bold=True)
            for row, doc in enumerate(docs, 2):
                for col, k in enumerate(keys, 1):
                    wd.cell(row=row, column=col, value=doc.get(k))
            for col in range(1, len(headers) + 1):
                wd.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 30

    elif isinstance(obj, list) and obj:
        ws = wb.create_sheet("Dados")
        headers = list(obj[0].keys())
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h).font = Font(bold=True)
        for row, item in enumerate(obj, 2):
            for col, k in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=str(item.get(k, "") or ""))

    wb.save(path)


def _normalizar_cnpj(cnpj: str) -> str:
    return "".join(c for c in cnpj if c.isdigit())


if __name__ == "__main__":
    sys.exit(main())
